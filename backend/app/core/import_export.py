"""Import/export Excel (xlsx) и CSV (US-4..US-6, UC-2, UC-4, FR-4..FR-6).

Безопасность (NFR-3/NFR-4):
- Excel читается ТОЛЬКО через openpyxl data_only=True (без формул, без макросов);
- нет execute вообще; значения не читаются из формул;
- санитизация имени файла при экспорте (Content-Disposition).
"""
from __future__ import annotations

import csv
import io
import os
import re
from typing import List, Optional, Tuple

from .plan import COLUMN_ALIASES, COLUMNS, PlanSchema, Task, canonical_predecessor_name


class ImportError_(Exception):
    """Ошибка формата/структуры входного файла (маппится в HTTP 400)."""

    def __init__(self, message: str):
        super().__init__(message)
        self.message = message


MAX_UPLOAD_BYTES = 20 * 1024 * 1024  # FR-4: ≤20 МБ
MAX_CELLS = 100_000  # защита от гигантских файлов (общая граница)


def _cell_value(raw) -> str:
    """Строковое представление значения ячейки (данные, не формулы)."""
    if raw is None:
        return ""
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return str(raw).strip()


def _normalize_text(values: List[Optional["object"]]) -> List[str]:
    """Приводим заголовок/строку к каноническим именам колонок через синонимы."""
    result: List[str] = []
    for h in values:
        key = _cell_value(h).strip().lower()
        result.append(COLUMN_ALIASES.get(key, ""))
    return result


def parse_xlsx_sheet_rows(data: bytes) -> Tuple[List[str], List[dict]]:
    """Парсит первый активный лист в (нормализованные headers, row-dicts).

    data: содержимое .xlsx. data_only=True — значения без формул/макросов.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    if not wb.sheetnames:
        raise ImportError_("Файл не содержит листов; добавьте хотя бы один лист с задачами")
    ws = wb[wb.sheetnames[0]]  # первый активный лист (roadmap: выбор листа)
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ImportError_("Лист пуст — нет строк данных")

    headers = _normalize_text(list(header_row))
    rows: List[dict] = []
    cell_count = 0
    for r in rows_iter:
        cell_count += len(r)
        if cell_count > MAX_CELLS:
            raise ImportError_("Файл слишком большой: больше 100 000 ячеек")
        if all(_cell_value(v) == "" for v in r):
            continue  # пропускаем пустые строки
        row = {}
        for i, h in enumerate(headers):
            if h:
                row[h] = _cell_value(r[i]) if i < len(r) else ""
        rows.append(row)
    wb.close()
    return headers, rows


def parse_csv_rows(data: bytes) -> Tuple[List[str], List[dict]]:
    """Парсит CSV (utf-8-sig для BOM, delimiter auto через sniffer-подобный выбор)."""
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4000]
    try:
        delimiter = csv.Sniffer().sniff(sample).delimiter
    except csv.Error:
        delimiter = ";"
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = _normalize_text(reader.fieldnames or [])
    rows = []
    n = 0
    for r in reader:
        row = {}
        for i, h in enumerate(headers):
            if h:
                keys = list(r.keys())
                val = r[keys[i]] if i < len(keys) else ""
                row[h] = _cell_value(val)
        rows.append(row)
        n += 1
        if n > MAX_CELLS:
            raise ImportError_("Файл слишком большой: больше 100 000 строк")
    return headers, rows


def parse_tasks(data: bytes, filename: str) -> Tuple[List[Task], List[str]]:
    """Человекопонятный парсинг файла в задачи.

    Returns: (tasks, warnings). Бросает ImportError_ для фатальных ошибок структуры.
    """
    base, ext = os.path.splitext(filename.lower())
    if ext == ".xlsx":
        headers, rows = parse_xlsx_sheet_rows(data)
    elif ext == ".csv":
        headers, rows = parse_csv_rows(data)
    else:
        raise ImportError_("Поддерживаются только .xlsx и .csv")

    missing = [c for c in COLUMNS if c not in headers]
    if missing:
        raise ImportError_("Отсутствуют колонки: " + ", ".join(missing))
    if not rows:
        raise ImportError_("Файл пуст — нет данных задач")

    index_by_name: dict[str, str] = {}
    tasks: List[Task] = []
    warnings: List[str] = []
    row_index = 2  # после заголовка (1-based для сообщений)

    for r in rows:
        row_index += 1
        name = r.get("задача", "")
        if not name:
            warnings.append(f"Строка {row_index}: пустое имя задачи — пропущена")
            continue
        if name in index_by_name:
            warnings.append(f"Строка {row_index}: дублируется задача «{name}» — оставлена последняя")

        duration_raw = r.get("длительность", "1")
        try:
            duration = max(1, int(float(duration_raw)))
        except (TypeError, ValueError):
            duration = 1
            warnings.append(f"Задача «{name}»: длительность '{duration_raw}' → 1 день")

        preds = canonical_predecessor_name(r.get("предшественники", ""))
        tasks.append(
            Task(
                name=name,
                description=r.get("описание", ""),
                assignee=r.get("исполнитель", ""),
                duration_days=duration,
                predecessors=preds,
                original_row=row_index,
            )
        )
        index_by_name[name] = name

    # Валидация предшественников: ссылки на несуществующие имена — предупреждение.
    known = set(index_by_name)
    for i, t in enumerate(tasks):
        unresolved = [p for p in t.predecessors if p not in known]
        if unresolved:
            warnings.append(
                f"Задача «{t.name}»: неизвестные предшественники {', '.join(unresolved)} — игнорированы"
            )
            t.predecessors = [p for p in t.predecessors if p in known]

    return tasks, warnings


# --- Экспорт -----------------------------------------------------------------

def _sanitize_filename(name: str) -> str:
    """Безопасное имя файла для Content-Disposition (FR-6)."""
    name = os.path.basename(name or "план")
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name).strip()
    return name or "план"


def export_tasks(tasks: List[Task], fmt: str = "xlsx", source: str = "план") -> Tuple[bytes, str]:
    """Экспорт плана в xlsx/csv. Возвращает (bytes, filename).

    Колонки: задача/описание/исполнитель/длительность/предшественники + начало/конец.
    """
    headers = COLUMNS + ["начало", "конец"]
    rows = []
    for t in sorted(tasks, key=lambda x: x.start_day or 0):
        rows.append(
            {
                "задача": t.name,
                "описание": t.description,
                "исполнитель": t.assignee,
                "длительность": t.duration_days,
                "предшественники": ", ".join(t.predecessors),
                "начало": t.start_day if t.start_day else "",
                "конец": t.end_day if t.end_day else "",
            }
        )

    base = _sanitize_filename(os.path.splitext(source)[0]) or "план"
    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
        return buf.getvalue().encode("utf-8-sig"), f"{base}.csv"
    if fmt == "xlsx":
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "план"
        ws.append(headers)
        for r in rows:
            ws.append([r[c] for c in headers])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), f"{base}.xlsx"
    raise ValueError("Формат экспорта должен быть xlsx или csv")


def schema_of(tasks: List[Task], source: str = "") -> PlanSchema:
    ends = [t.end_day for t in tasks if t.end_day is not None]
    return PlanSchema(
        n_tasks=len(tasks),
        total_days=max(ends) if ends else 0,
        source_filename=source,
    )